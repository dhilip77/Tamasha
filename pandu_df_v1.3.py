# -*- coding: utf-8 -*-
"""
Created on Fri May  4 10:29:47 2018

@author: Dhilip.KumarTG
"""
import sys
sys.path.append('C:\Program Files\Python36\openpyxl-2.5.3\jdcal-1.4')
sys.path.append('C:\Program Files\Python36\Lib\pandas-0.23.4\pandas-0.23.4')
#sys.path.append('C:\Program Files\Python36\Lib\numpy-1.15.2')
sys.path.append('C:\Program Files\Python36\Lib\numpy-1.15.2')
sys.path.append('C:\Program Files\Python36\Lib\pytz-2018.5')
sys.path.append('C:\Program Files\Python36\Lib\python-dateutil-2.7.3')
import openpyxl as pyxl
import datetime
import pandas as pd
#from openpyxl.chart import LineChart,Reference
"""Note: THREE PATH LOCATIONS TO BE CONSIDER AFTER MOVING CODE TO DIFFERENT MACHINE"""
"""Main caller function does try and catch"""

def loadWorkbook():
    try:
        xlFile,iFile = fileRead()       
        wbk = pyxl.load_workbook(xlFile)
        sheets = wbk.sheetnames
        print(sheets)
        print("Sheets has to match Tables Summary NONPRD Alerts")
       
        NAlertSheet = wbk["NONPRD Alerts"]
        TabSheet = wbk["Tables"]
        SumSheet = wbk["Summary"]       
        #print(TabSheet.cell(row=2, column=12).value)        
        NARow = NAlertSheet.max_row
        iDates = iDateCount(NARow, NAlertSheet)
        idict = iAlertCount(iDates,NARow,NAlertSheet)
        appSet = iAppCount(NARow,NAlertSheet)
        temp = appAlertCount(appSet,NARow,NAlertSheet)
        WriteTable(idict,TabSheet,NARow)   
        WriteAppTable(temp,TabSheet)
        jSet= JobUniqCount(NARow,NAlertSheet)
        jtemp,aLertcnt= JobAlertCount(jSet ,NARow, NAlertSheet)
        WriteJobTable(jtemp,TabSheet)        
        try:
            files = fileSave(idict)
            Is_Duplicate_Afiles(idict,aLertcnt-1,TabSheet,wbk,files)
        except IOError:
            print("output dir is not writable--->>")
    except IOError:
        print("Error at Excel path!!!")
        print("Check your new.txt file for PATH!!!")
    except ImportError:
        print("Error on file import!!!")
    finally:
        wbk.save(files)
        ChartsInit(TabSheet,SumSheet,wbk,files)
        #Chart_test(files)
        print("END File Closed !!!")
        print("Chart preparing...SAVED at Others DIR....",files)
        
        wbk.close()
        iFile.close()
        
def fileSave(idict):
    recent_date = list(idict.keys())
    file_date = str(recent_date[1])
    format_str = '%d/%m/%Y'
    file_date = str(datetime.datetime.strptime(file_date, format_str).strftime("%Y-%m-%d"))
    files = "C:\\Users\dhilip.kumartg\\Documents\\Alerts_report\\report\\ControlM_NONProd_Offenders_"+file_date+".xlsx"
    return files    
    
"""Read Notepad file has link to excel sheet"""
def fileRead():
    iFile = open("C:\\Users\dhilip.kumartg\\nonPrdOffenders\\src\\new.txt", 'r')
    fileRead.datum = iFile.readline()
    return fileRead.datum,iFile


"""For each Job count total alerts count"""
def JobAlertCount(jSet ,NARow, NAlert):
    idict = {}
    for y in jSet:
        cnt = 0
        for i in range(1, NARow):
            ix = NAlert.cell(row=i, column=5).value
            if y == ix:
                #print("ix:",ix)
                cnt += 1
            else:
                cnt = cnt
        idict[y] = cnt
        #print(y,"=",cnt)
    print("Total Job Alerts::",NARow-2)
    #print(idict)
    jtemp = [(k, idict[k]) for k in sorted(idict, key=idict.get, reverse = True)]
    #print("App_TEMP=",temp)
    #for k, v in temp:
        #print(k, v) 
    #print(jtemp)
    return jtemp,NARow


""" Total alerts count as per each date"""
def iAlertCount(iDates,NARow,NAlert):
    idict = {}
    for y in iDates:
        cnt = 0
        for i in range(1, NARow + 1):
            ix = NAlert.cell(row=i, column=1).value
            if ix == "HOST_TIME":
                cnt = cnt
                #print(ix)
            elif y == ix:
                #print("ix:",ix)
                cnt = cnt + 1
            else:
                cnt = cnt
        idict[y] = cnt
        #print(y,"=",cnt)
    print("Total Alerts per date==",NARow-1)
    #print(idict)
    ytemp = [(k, idict[k]) for k in sorted(idict, reverse = True)]
    idict = dict(ytemp)
    return idict


"""Total alerts count as per Application"""
def appAlertCount(appSet,NARow, NAlert):
    idict = {}
    for y in appSet:
        cnt = 0
        for i in range(1, NARow):
            ix = NAlert.cell(row=i, column=3).value
            if y == ix:
                #print("ix:",ix)
                cnt += 1
            else:
                cnt = cnt
        idict[y] = cnt
        #print(y,"=",cnt)
    print("Total Alerts=",NARow-2)
    #print(idict)
    temp = [(k, idict[k]) for k in sorted(idict, key=idict.get, reverse = True)]
    #print("App_TEMP=",temp)
    #for k, v in temp:
        #print(k, v)    
    return temp


"""Count total unique jobs List"""
def JobUniqCount(NARow,NAlert):
    jSet = set()
    for ai in range(1, NARow):
        ix = NAlert.cell(row=ai, column=5).value
        jSet.add(ix)
    #print("SET:",jSet)
    return jSet

   
"""Total Unique Applications List"""
def iAppCount(NARow,NAlert):
    aSet = set()
    for ai in range(1, NARow):
        ix = NAlert.cell(row=ai, column=3).value
        aSet.add(ix)
    #print("SET:",aSet)
    return aSet
    

"""Total Unique Date List"""  
def iDateCount(NARow, NAlert):
    dSet = set()
    for itr in range(1, NARow):
        ix = NAlert.cell(row=itr, column=1).value
        dSet.add(ix)  
    #print(dSet)
    return dSet    


"""Write to excel Table has Date and total alerts count"""
def WriteTable(idict,TabSheet,NARow):
    rNum = 2
    #print(idict)
    TabSheet.cell(row=1, column=12).value = "Date"
    TabSheet.cell(row=1, column=13).value = "Total Alerts"
    for key, val in idict.items():
        if key == "HOST_TIME":
            print("Host time:")
            rNum -= 1
        else:    
            TabSheet.cell(row=rNum, column=12, value = key)
            TabSheet.cell(row=rNum, column=13, value = val)
        rNum += 1
    TabSheet.cell(row=rNum, column=12).value = "Grand Total"
    TabSheet.cell(row=rNum, column=13).value = "=SUM(M2:M8)"
    

"""Write to excel has Application and total alerts count"""    
def WriteAppTable(temp,TabSheet):
    AlertDt = {}
    AlertDt = dict(temp)
    rNum = 2
    #print("dc=", AlertDt)
    for key, val in AlertDt.items():
        TabSheet.cell(row=rNum, column=9, value= key)
        TabSheet.cell(row=rNum, column=10, value= val)
        rNum += 1
    TabSheet.cell(row=1, column=9).value = "APPLICATION"
    TabSheet.cell(row=1, column=10).value ="ALERTS"


"""Write table to excel has total Job and alerts count"""
def WriteJobTable(jtemp,TabSheet):
    AlertDt = {}
    AlertDt = dict(jtemp)
    rNum = 2
    #print("dc=", AlertDt)
    for key, val in AlertDt.items():
        TabSheet.cell(row=rNum, column=15, value= key)
        TabSheet.cell(row=rNum, column=16, value= val)
        rNum += 1
    TabSheet.cell(row=1, column=15).value = "JOBs"
    TabSheet.cell(row=1, column=16).value ="ALERTS"
    
""" Charts on Summary Sheet
def ChartsInit(files,cwbk):
    #cwb = pyxl.load_workbook(files)
    csheets = cwbk.sheetnames
    print(csheets)
    ctabsheet = cwbk["Tables"]
    csumsheet = cwbk["Summary"]
    
    TenOffenders(ctabsheet,csumsheet)
    #SerChart(ctabsheet,csumsheet)
    AlertsDate(ctabsheet,csumsheet)
    #cwbk.save(files)
    #return csumsheet,ctabsheet,cwbk   

"""
    
""" Charts on Summary Sheet"""
def ChartsInit(TabSheet,SumSheet,wbk,files):
    TenOffenders(TabSheet,SumSheet)
    ChartAlertsDate(TabSheet,SumSheet,wbk,files)
    #ChartVarian(TabSheet,SumSheet,wbk,files)

"""Draw a Bar chart"""
def ChartAlertsDate(ctabsheet,csumsheet,wbk,files):
    dateCrt = pyxl.chart.BarChart()
    dateCrt.type = "col"
    dateCrt.sytle = 10
    dateCrt.title = "Alert per Date"
    dateCrt.y_axis.title = "Alerts in Numbers"
    dateCrt.x_axis.title = "Alerts per Date"
    
    Ddata = pyxl.chart.Reference(ctabsheet, min_col= 12, min_row=2, max_col=13, max_row=8)
    dateCrt.add_data(Ddata, titles_from_data=False)
    dateCrt.set_categories(Ddata)
    dateCrt.shape=4
    csumsheet.add_chart(dateCrt, "A3")
    wbk.save(files)
    wbk.close()
    

"""Table of top twenty offenders Range increased from 12 to 22"""
def TenOffenders(ctabsheet, csumsheet):
    xdict = {}
    for cnt in range(1,22):
        ky = ctabsheet.cell(row=cnt, column =15).value
        vlu = ctabsheet.cell(row=cnt, column =16).value
        xdict[ky] = vlu
    #print(xdict)
    rcnt = 22
    for x, y in xdict.items():
        csumsheet.cell(row=rcnt, column = 4).value = x
        csumsheet.cell(row=rcnt, column = 5).value = y
        rcnt += 1
    csumsheet.cell(row=20, column = 4).value = "TOP TEN OFFENDERS"

"""The New excel file data is check with Archive sheet"""
"""The date is extracted from new excel sheet"""
def Is_Duplicate_Nfiles(idict):
    dSet = set()
    for key, val in idict.items():
        kye = key
        if kye != 'HOST_TIME':
            #print("Date Key:",kye)
            dSet.add(kye)
        else:
            print("host_time")
    #print("Date Set:",dSet) 
    return dSet

"""To read Archive File Sheet"""
def ArchFileRead():
    try:
        fPath = "C:\\Users\dhilip.kumartg\\nonPrdOffenders\\src\\arc.xlsx"
        awbk = pyxl.load_workbook(fPath)
        ASheet = awbk.sheetnames
        print("Archive Sheet Available...-->",ASheet)
    except IOError:
        print("Check the code-PATH for Archieve File is Missing at misc\\arc.xlsx ")
        exit(1)
    finally:
       return awbk,fPath
   
"""To get a Date set from Archieve Sheet """
def ArchDateSet(awbk):
    dList = []
    cnt = 2
    Asheet = awbk['Arch']
    ArchMaxRow = Asheet.max_row
    #print("Arch Sheet Max rows:",ArchMaxRow)
    for itr in range(1,ArchMaxRow):
        dset = Asheet.cell(row=cnt, column=2).value
        #print(dset)
        if dset == "NONE":
            print("ERROR AT ARCHIVE SHEET 'ARCH' NONE OR EMPTY ROW WAS UPDATED")
            print("CALCULATIONS ARE NOT PERFORMED, CHECK ARCH SHEET FOR EMPTY ROW")
        else:
            dList.insert(itr,dset)
        cnt += 1
    #print(dList)
    return dList, Asheet

"""Comparison operator"""
def Comp_dates(dList,dSet):
    ASet = ()
    ASet = set(dList)
    #print("+++++",ASet,"++++",dSet)
    if ASet.intersection(dSet):
        print("TRUE-- OLD FILE DATA...")
        print("REMOVE LIST DATE FROM FILE..")
        print(ASet & dSet)
        return True
    else:
        print("FALSE-- NEW FILE DATA...")
        print(ASet & dSet)
        return False

"""Check for duplicate date in Archive file"""
"""BOOLEAN FUNCTION TO UPDATE ARCHIVE FILE"""
def Is_Duplicate_Afiles(idict,AlertCnt,TabSheet,wbk,files):
    dSet = Is_Duplicate_Nfiles(idict)
    awbk,AfPath = ArchFileRead()
    #print(awbk)
    dList, Asheet = ArchDateSet(awbk)
    Val_bool = Comp_dates(dList,dSet)
    #print(dList)
    print(Val_bool)
    """Check for Old data file"""
    if Val_bool == False:
        Update_Archive_Sheet(Asheet,idict,AlertCnt)
        awbk.save(AfPath)
        Mig_Arch_Summary(TabSheet,Asheet,AfPath)
        wbk.save(files)
        wbk.close()
    else:
        print("ERROR UPDATE ON ARCIVE FILE SHEET, OLD DATA OR OLD FILE USED")
        Mig_Arch_Summary(TabSheet,Asheet,AfPath)
        #exit(2)

""" To update Archive Sheet with new data"""
def Update_Archive_Sheet(Asheet,idict,AlertCnt):
    #print(idict)
    AList = list(idict.keys())
    #print("Date String:", AList[1])
    """To get the recent date from sheet"""
    date_str = str(AList[1])
    format_str = '%d/%m/%Y'
    datetime_obj = datetime.datetime.strptime(date_str, format_str)
    #print(datetime_obj)       
    """To get Week number"""
    weeknumber = datetime_obj.isocalendar()[1]
    print("WEEK OF YEAR:",weeknumber)
    """Read the Archive workbook"""
    #sheets = Archwrkbuk.sheetnames
    #print("Archive File Sheet:",sheets)
    ArchSheet = Asheet
    Row_max = ArchSheet.max_row
    print("Rows in ARCHIVE Sheet--->",Row_max)
    """To get the next successive row in Arch Sheet"""
    Row_next = Row_max + 1
    Var_val, Diff_val = ArchSheet_Calculation(ArchSheet,AlertCnt,Row_max)
    """ Update Sheet"""
    ArchSheet.cell(row = Row_next, column = 1).value = weeknumber
    ArchSheet.cell(row = Row_next, column = 2).value = date_str
    ArchSheet.cell(row = Row_next, column = 3).value = AlertCnt
    ArchSheet.cell(row = Row_next, column = 4).value = Var_val
    ArchSheet.cell(row = Row_next, column = 5).value = Diff_val
    ArchSheet.cell(row = Row_next, column = 4).number_format = '00.00%'
    ArchSheet.cell(row = Row_next, column = 5).number_format = '00.00%'
    
    
"""To do Calculations of Variation and Difference with current and old data"""    
def ArchSheet_Calculation(ArchSheet,AlertCnt,Row_max):
    Prev_column_Alert = ArchSheet.cell(row = Row_max, column = 3).value
    Prev_column_Var = ArchSheet.cell(row = Row_max, column = 4).value
    #print("Previous column Alert:", Prev_column_Alert, Prev_column_Var)
    Var_val = ((int(Prev_column_Alert)-int(AlertCnt))/Prev_column_Alert)
    Var_val = abs(float(Var_val))
    print("Previous Alert:", Prev_column_Alert,"Previous Variance:",Prev_column_Var)
    print("Variation after Format:", Var_val)
    Prev_column_Var = float(Prev_column_Var)
    Diff_val = float(Var_val - Prev_column_Var)
    Diff_val = float(abs(Diff_val))
    return Var_val, Diff_val

def Sheet_header(TabSheet):
    TabSheet.cell(row=1, column=1).value = "Week"
    TabSheet.cell(row=1, column=2).value = "Sunday"
    TabSheet.cell(row=1, column=3).value = "NONPrd Alerts"
    TabSheet.cell(row=1, column=4).value = "Variation"
    TabSheet.cell(row=1, column=5).value = "Difference"

def Sheet_Update(TabSheet,idx,weekday,Sunday,NAlerts,Varian,Differ):
    TabSheet.cell(row=idx, column=1).value = weekday
    TabSheet.cell(row=idx, column=2).value = Sunday
    TabSheet.cell(row=idx, column=3).value = NAlerts
    TabSheet.cell(row=idx, column=4).value = Varian
    TabSheet.cell(row=idx, column=5).value = Differ
    #TabSheet.cell(row=idx, column=4).number_format = style.numbers.FORMAT_PERCENTAGE
    #TabSheet.cell(row=idx, column=5).number_format = style.numbers.FORMAT_PERCENTAGE  
    
    
"""To Do migration of Old data to Tables Sheet at workbook"""
def Mig_Arch_Summary(TabSheet,Archsheet,AfPath):
    data_frame = pd.read_excel(AfPath, sheet_name='Arch')
    """READ ARCH IN DATAFRAME"""
    #print(data_frame)
    Sheet_header(TabSheet)
    Max_Row = Archsheet.max_row
    weekday = data_frame.iloc[-1]["Week"]
    #print("Archive Weekday:",weekday)
    for idx in range(1, Max_Row):
        if(idx <= Max_Row -1):
            weekday = data_frame.iloc[-idx]["Week"]
            Sunday = data_frame.iloc[-idx]["Sunday"]
            NAlerts = data_frame.iloc[-idx]["NONPrd Alerts"]
            Varian = data_frame.iloc[-idx]["Variation"]
            Differ = data_frame.iloc[-idx]["Difference"]
            Varian = float(abs(Varian % 100))
            Differ = float(abs(Differ % 100))
            #print("Differ:",Differ)
            #print("Varian:",Varian)
            Sheet_Update(TabSheet,idx + 1,weekday,Sunday,NAlerts,Varian,Differ)
        elif(idx == Max_Row):
            break
        
        
"""Variation Chart      
def Chart_test(files):
    wkb = pyxl.load_workbook(files)
    sheet = wkb["Tables"]
    values = Reference(sheet, min_col = 4, min_row = 2, 
                         max_col = 4, max_row = 4)  
# Create object of LineChart class 
    chart = LineChart()   
    chart.add_data(values)  
# set the title of the chart 
    chart.title = " Variation-CHART " 
# set the title of the x-axis 
    chart.x_axis.title = " Weekly "  
# set the title of the y-axis 
    chart.y_axis.title = " Percentage "
# add chart to the sheet 
# the top-left corner of a chart 
# is anchored to cell E2 . 
    sheet.add_chart(chart, "A25")
# save the file 
    wkb.save(files)
    wkb.close()

"""  
loadWorkbook()





